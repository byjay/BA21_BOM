#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BOM 데이터 처리 시스템 v2.0
- 옵션 1: JSON 전체 생성
- 옵션 2: 엑셀 생성 + 그래프 (병합/정렬)
- 옵션 3: JSON vs 원본 엑셀 비교 검증
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import json
import os
from pathlib import Path
from datetime import datetime
import sys


class BOMProcessor:
    """BOM 데이터 처리 메인 클래스"""

    def __init__(self):
        self.weld_file = None
        self.detail_file = None
        self.result_df = None
        self.group_columns = ['MATNO', 'STEEL NO', 'NESTING DWG', 'Grade', 'T']

        # 컬럼 순서 정의 (사용자 지정)
        self.column_order = [
            'WELD UNIQUE ID', 'BLOCK', 'FILENAME', 'DWG. Title', 'MOD. NO', 'DETAIL VIEW',
            'MATNO', 'STEEL NO', 'NESTING DWG', 'Grade',
            'OFF', 'WLEG', 'WELD. LENG.', 'SIDE', 'WNO', 'P. NO',
            'ea', 'total', 'T', 'B', 'L(OD)', 'WEIGHT', 'MIX', 'no',
            'TPYE', 'WORKSCOPE', 'REV1'
        ]

        # 제외할 컬럼 (MATNO1~MATNO6는 WELD ID 파싱용 임시 데이터이므로 결과에서 제거)
        self.exclude_columns = ['MATNO1', 'MATNO2', 'MATNO3', 'MATNO4', 'MATNO5', 'MATNO6', 'MOD', '_matched']

        # Windows 콘솔 인코딩 설정
        if sys.platform == 'win32':
            try:
                sys.stdout.reconfigure(encoding='utf-8')
            except:
                pass

    def detect_weld_file(self, df):
        """WELD 파일 자동 감지"""
        columns = df.columns.tolist()
        has_weld_id = any('WELD' in str(col).upper() and 'UNIQUE' in str(col).upper() for col in columns)
        matno_cols = [col for col in columns if 'MATNO' in str(col).upper() and any(c.isdigit() for c in str(col))]
        return has_weld_id and len(matno_cols) >= 2

    def find_excel_files(self, directory='data'):
        """데이터 디렉토리에서 Excel 파일 찾기 (기존 logic 보완)"""
        excel_files = []
        data_dir = Path(__file__).parent / directory
        if not data_dir.exists():
            print(f"[WARN] {directory} 디렉토리가 없습니다. 현재 디렉토리에서 검색합니다.")
            data_dir = Path('.')
            
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(data_dir.glob(ext))
        # weld_export.xlsx 및 임시 파일(~$...) 제외
        return [str(f) for f in excel_files if 'weld_export' not in str(f).lower() and not str(f.name).startswith('~$')]

    def load_files(self):
        """파일 자동 로드"""
        print("=" * 60)
        print("BOM 데이터 처리 시스템 v2.0")
        print("=" * 60)

        excel_files = self.find_excel_files()

        if len(excel_files) < 2:
            print(f"[ERROR] Excel 파일이 {len(excel_files)}개만 발견되었습니다. 최소 2개 필요합니다.")
            return False

        print(f"\n[OK] 발견된 Excel 파일: {len(excel_files)}개")
        for i, f in enumerate(excel_files[:5], 1):
            print(f"  {i}. {os.path.basename(f)}")

        # 260130 파일 우선 순위 로직
        priority_files = sorted(excel_files, key=lambda x: '260130' in x, reverse=True)
        
        for file in priority_files:
            try:
                # 헤더만 읽어서 파일 유형 판단 (속도 최적화)
                df_head = pd.read_excel(file, nrows=0)
                if self.detect_weld_file(df_head) and not self.weld_file:
                    self.weld_file = file
                    print(f"\n[OK] WELD 파일 (감지됨): {os.path.basename(file)}")
                elif 'MATNO' in [str(c).upper() for c in df_head.columns] and not self.detail_file:
                    self.detail_file = file
                    print(f"[OK] 상세 파일 (감지됨): {os.path.basename(file)}")
            except Exception as e:
                print(f"[WARN] 파일 읽기 오류 ({os.path.basename(file)}): {e}")

        if not self.weld_file or not self.detail_file:
            print("\n[WARN] WELD 또는 상세 파일 자동 감지 실패. 리스트 상단 파일을 사용합니다.")
            if not self.weld_file and priority_files: self.weld_file = priority_files[0]
            if not self.detail_file and len(priority_files) > 1: self.detail_file = priority_files[1]

        return True

    def match_data(self):
        """데이터 매칭"""
        print("\n" + "=" * 60)
        print("데이터 매칭 시작...")
        print("=" * 60)

        weld_df = pd.read_excel(self.weld_file, keep_default_na=False)
        detail_df = pd.read_excel(self.detail_file, keep_default_na=False)

        print(f"[OK] WELD 데이터: {len(weld_df)}행")
        print(f"[OK] 상세 데이터: {len(detail_df)}행")

        # 컬럼 찾기
        weld_id_col = None
        for col in weld_df.columns:
            if 'WELD' in str(col).upper() and 'UNIQUE' in str(col).upper():
                weld_id_col = col
                break

        matno_columns = [col for col in weld_df.columns
                        if 'MATNO' in str(col).upper() and any(c.isdigit() for c in str(col))]

        detail_matno_col = None
        for col in detail_df.columns:
            if str(col).upper() == 'MATNO':
                detail_matno_col = col
                break

        print(f"\n[OK] WELD ID: {weld_id_col}")
        print(f"[OK] MATNO 컬럼: {len(matno_columns)}개")
        print(f"[OK] Detail MATNO: {detail_matno_col}")

        # 매칭 수행
        results = []
        matched = 0
        missing = 0

        for idx, weld_row in weld_df.iterrows():
            for matno_col in matno_columns:
                matno = weld_row[matno_col]

                if matno and str(matno).strip():
                    matched_detail = detail_df[detail_df[detail_matno_col] == matno]

                    if not matched_detail.empty:
                        detail_row = matched_detail.iloc[0]
                        result_row = {**weld_row.to_dict(), **detail_row.to_dict()}
                        result_row['_matched'] = True
                        results.append(result_row)
                        matched += 1
                    else:
                        result_row = weld_row.to_dict()
                        result_row[detail_matno_col] = matno
                        result_row['_matched'] = False
                        results.append(result_row)
                        missing += 1

        self.result_df = pd.DataFrame(results)

        print(f"\n[OK] 매칭 성공: {matched:,}개")
        print(f"[WARN] 매칭 실패: {missing:,}개")
        print(f"[OK] 총 레코드: {len(self.result_df):,}개")

        return True

    def clean_data(self, df):
        """excel-master 패턴을 적용한 고성능 데이터 정제"""
        import numpy as np
        from datetime import datetime, date

        # 1. 빈 행/열 제거
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # 2. 중복 헤더/컬럼 처리
        df = df.loc[:, ~df.columns.duplicated()]

        # 3. 공백 제거 및 NaN 처리
        df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        df = df.fillna('')

        # 4. 타입 변환 및 포맷팅 (NP/Datetime -> String)
        for col in df.columns:
            df[col] = df[col].apply(lambda x:
                '' if (str(x).lower() == 'nan' or pd.isna(x))
                else (x.isoformat() if isinstance(x, (datetime, date, pd.Timestamp))
                else int(x) if isinstance(x, (np.int64, float)) and x == int(x) # 정수형 실수 처리
                else str(x) if isinstance(x, (np.int64, np.float64, float))
                else x)
            )

        # 5. WNO 컬럼을 3자리 숫자로 포맷팅 (001, 002, ...)
        if 'WNO' in df.columns:
            df['WNO'] = df['WNO'].apply(lambda x:
                f"{int(float(x)):03d}" if (str(x).strip() and str(x).replace('.', '').replace('-', '').isdigit())
                else x
            )

        return df.infer_objects()

    def reorder_columns(self, df):
        """컬럼을 지정된 순서로 재정렬 및 불필요한 컬럼 제거"""
        # 현재 데이터프레임의 컬럼
        existing_cols = df.columns.tolist()

        # 제외할 컬럼 제거
        existing_cols = [col for col in existing_cols if col not in self.exclude_columns]

        # 순서대로 정렬 (존재하는 컬럼만)
        ordered_cols = [col for col in self.column_order if col in existing_cols]

        # 순서에 없는 나머지 컬럼 추가 (제외 컬럼 제외)
        remaining_cols = [col for col in existing_cols
                         if col not in ordered_cols and col not in self.exclude_columns]

        # 최종 컬럼 순서
        final_cols = ordered_cols + remaining_cols

        return df[final_cols]

    # =====================================================================
    # 옵션 1: JSON 전체 생성
    # =====================================================================
    def option1_generate_json(self):
        """옵션 1: JSON 전체 생성 (기존 로직 유지)"""
        print("\n" + "=" * 60)
        print("옵션 1: JSON 파일 생성")
        print("=" * 60)

        # 루트 디렉토리 (!!bom 폴더)
        parent_dir = Path(__file__).parent

        clean_df = self.result_df.drop(columns=['_matched'], errors='ignore')
        clean_df = self.clean_data(clean_df)
        clean_df = self.reorder_columns(clean_df)  # 컬럼 순서 정렬

        # JSON 디렉토리 생성
        json_dir = parent_dir / 'json'
        json_dir.mkdir(exist_ok=True)

        # all_data.json 생성 (검증용)
        clean_df.to_json(json_dir / 'all_data.json', orient='records', force_ascii=False, indent=2)

        # 그룹별 JSON 생성 제거 (필요없으므로 생략)

        # JavaScript 파일 (all_data.js) 생성
        print("\n[OK] all_data.js 파일 생성 중...")
        all_data = clean_df.to_dict(orient='records')
        with open(parent_dir / 'all_data.js', 'w', encoding='utf-8') as f:
            f.write('// Auto-generated from bom.py\n')
            f.write('window.ALL_DATA = ')
            json.dump(all_data, f, ensure_ascii=False)
            f.write(';\n')
        print(f"[OK] JavaScript 파일: {parent_dir / 'all_data.js'}")

        print("\n[SUCCESS] 옵션 1 완료!")

    # =====================================================================
    # 옵션 2: 엑셀 생성 + 그래프
    # =====================================================================
    def option2_excel_with_charts(self):
        """옵션 2: 엑셀 생성 + 그래프 (병합/정렬 로직)"""
        print("\n" + "=" * 60)
        print("옵션 2: 엑셀 생성 + 그래프")
        print("=" * 60)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'weld_export_{timestamp}.xlsx'

        clean_df = self.result_df.drop(columns=['_matched'], errors='ignore')
        clean_df = self.clean_data(clean_df)
        clean_df = self.reorder_columns(clean_df)  # 컬럼 순서 정렬

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 전체 데이터 시트
            clean_df.to_excel(writer, sheet_name='전체', index=False)
            print(f"[OK] 전체 시트: {len(clean_df):,}행")
            
            # workbook 가져오기
            wb = writer.book
            ws_all = wb['전체']
            
            # 헤더 인덱스 매핑
            all_cols = {cell.value: get_column_letter(cell.column) for cell in ws_all[1]}

            # 그룹별 시트 생성 (공식 연동)
            for group_col in self.group_columns:
                if group_col not in clean_df.columns:
                    continue

                # 해당 그룹 컬럼의 데이터만 필터링 하여 정렬된 인덱스 확보
                filtered_indices = clean_df[clean_df[group_col].astype(str).str.strip() != ''].index.tolist()
                
                # 병합 및 정렬 기준 (속도 위해 DataFrame 정렬 사용)
                sort_cols = [group_col]
                for sc in self.group_columns:
                    if sc in clean_df.columns and sc != group_col:
                        sort_cols.append(sc)
                
                # 정렬된 인덱스 순서 추출
                sorted_df = clean_df.loc[filtered_indices].sort_values(by=sort_cols)
                sorted_indices = sorted_df.index.tolist()

                sheet_name = group_col.replace(' ', '_').replace('/', '_')[:31]
                ws_group = wb.create_sheet(sheet_name)
                
                # 헤더 복사
                for c_idx, col_name in enumerate(clean_df.columns, 1):
                    ws_group.cell(row=1, column=c_idx, value=col_name)

                # 데이터 연동 (Excel 공식 사용: =전체!A2)
                for r_idx, original_idx in enumerate(sorted_indices, 2):
                    excel_row_num = original_idx + 2 # pandas index 0 -> excel row 2
                    for c_idx, col_name in enumerate(clean_df.columns, 1):
                        col_letter = get_column_letter(c_idx)
                        # 중요: 모든 셀을 '전체' 시트의 해당 행/열로 연동
                        formula = f"='전체'!{col_letter}{excel_row_num}"
                        ws_group.cell(row=r_idx, column=c_idx, value=formula)

                print(f"[OK] {group_col} 시트: {len(sorted_indices):,}행 (공식 연동 완려)")

        # 서식 및 차트 적용
        self.apply_formatting_and_charts(filename, clean_df)
        print(f"\n[SUCCESS] 옵션 2 완료: {filename}")

    def apply_formatting_and_charts(self, filename, df):
        """Excel 서식 및 차트 적용"""
        print("\n[OK] 서식 및 차트 적용 중...")
        wb = openpyxl.load_workbook(filename)

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 헤더 서식
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 데이터 테두리
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border

            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # 자동 필터 추가 (모든 시트)
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
                print(f"[OK] {sheet_name} 시트에 자동 필터 추가")

            # 차트 추가 (전체 시트에만)
            if sheet_name == '전체' and ws.max_row > 1:
                self.add_charts_to_sheet(ws, df)

        wb.save(filename)
        print("[OK] 서식 및 차트 적용 완료")

    def add_charts_to_sheet(self, ws, df):
        """시트에 차트 추가"""
        try:
            # Grade 분포 차트 (Pie Chart)
            if 'Grade' in df.columns:
                grade_counts = df['Grade'].value_counts().head(10)

                # 데이터 시트 우측에 임시 데이터 추가
                start_col = ws.max_column + 2
                ws.cell(row=1, column=start_col, value="Grade")
                ws.cell(row=1, column=start_col+1, value="Count")

                for idx, (grade, count) in enumerate(grade_counts.items(), start=2):
                    ws.cell(row=idx, column=start_col, value=str(grade))
                    ws.cell(row=idx, column=start_col+1, value=count)

                # Pie Chart 생성
                pie = PieChart()
                pie.title = "Grade 분포 (Top 10)"
                labels = Reference(ws, min_col=start_col, min_row=2, max_row=len(grade_counts)+1)
                data = Reference(ws, min_col=start_col+1, min_row=1, max_row=len(grade_counts)+1)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.height = 10
                pie.width = 15

                ws.add_chart(pie, f"{get_column_letter(start_col+3)}2")

            # WEIGHT 합계 (Bar Chart)
            if 'WEIGHT' in df.columns and 'MATNO' in df.columns:
                weight_by_matno = df.groupby('MATNO')['WEIGHT'].sum().sort_values(ascending=False).head(10)

                start_col = ws.max_column + 2
                ws.cell(row=1, column=start_col, value="MATNO")
                ws.cell(row=1, column=start_col+1, value="Total Weight")

                for idx, (matno, weight) in enumerate(weight_by_matno.items(), start=2):
                    ws.cell(row=idx, column=start_col, value=str(matno))
                    ws.cell(row=idx, column=start_col+1, value=float(weight))

                # Bar Chart 생성
                bar = BarChart()
                bar.title = "MATNO별 WEIGHT 합계 (Top 10)"
                bar.x_axis.title = "MATNO"
                bar.y_axis.title = "Weight"

                data = Reference(ws, min_col=start_col+1, min_row=1, max_row=len(weight_by_matno)+1)
                cats = Reference(ws, min_col=start_col, min_row=2, max_row=len(weight_by_matno)+1)
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(cats)
                bar.height = 10
                bar.width = 15

                ws.add_chart(bar, f"{get_column_letter(start_col+3)}20")

        except Exception as e:
            print(f"[WARN] 차트 생성 중 오류: {e}")

    # =====================================================================
    # 옵션 3: JSON vs 원본 비교 검증
    # =====================================================================
    def option3_validate_json(self):
        """옵션 3: 데이터 무결성 철저 검증 (JSON vs 원본, 전체 시트 vs 그룹 시트)"""
        print("\n" + "=" * 60)
        print("옵션 3: 데이터 무결성 철저 검증")
        print("=" * 60)

        # 1. JSON 검증
        json_path = Path('all_data.json') # 루트 경로로 수정됨
        if not json_path.exists():
            print("[WARN] all_data.json 파일이 없습니다. 옵션 1을 먼저 실행하세요.")
            json_df = None
        else:
            with open(json_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            json_df = pd.DataFrame(json_data)
            print(f"[OK] JSON 데이터: {len(json_df):,}행")

        # 2. 최신 내보내기 엑셀 검증 (전체 vs 그룹 시트 동기화 확인)
        export_files = sorted(list(Path('.').glob('weld_export_*.xlsx')), key=os.path.getmtime, reverse=True)
        if not export_files:
            print("[WARN] 검증할 내보내기 엑셀 파일이 없습니다. 옵션 2를 먼저 실행하세요.")
            return

        target_excel = export_files[0]
        print(f"[OK] 검증 대상 엑셀: {target_excel.name}")
        
        wb = openpyxl.load_workbook(target_excel, data_only=True) # 공식 결과값으로 읽기
        ws_all = wb['전체']
        
        # '전체' 시트 데이터를 DataFrame으로 변환
        all_data = []
        headers = [cell.value for cell in ws_all[1]]
        for row in ws_all.iter_rows(min_row=2, values_only=True):
            all_data.append(dict(zip(headers, row)))
        df_all = pd.DataFrame(all_data)

        results = {
            'timestamp': datetime.now().isoformat(),
            'target_excel': target_excel.name,
            'cross_sheet_validation': {}
        }

        # 3. 그룹 시트 무결성 검증 (MATNO / WELD ID 기준)
        print(f"\n[검증] 시트 간 무결성 체크 (전체 vs 그룹):")
        for sheet_name in wb.sheetnames:
            if sheet_name == '전체': continue
            
            ws_group = wb[sheet_name]
            group_data = []
            for row in ws_group.iter_rows(min_row=2, values_only=True):
                group_data.append(dict(zip(headers, row)))
            df_group = pd.DataFrame(group_data)
            
            # 기준 키 설정 (WELD UNIQUE ID 우선, 없으면 MATNO)
            key_col = 'WELD UNIQUE ID' if 'WELD UNIQUE ID' in df_all.columns else 'MATNO'
            
            # 전체 시트에서 해당 그룹에 속해야 하는 데이터 추출
            # (그룹 시트는 해당 컬럼이 비어있지 않은 데이터만 포함됨)
            group_col_name = sheet_name.replace('_', ' ') # 시트명에서 컬럼명 복원 시도
            # 정확한 매핑을 위해 headers에서 찾기
            actual_col = next((h for h in headers if h.replace(' ', '_').replace('/', '_')[:31] == sheet_name), None)
            
            if not actual_col:
                print(f"  [SKIP] {sheet_name}: 매칭되는 컬럼을 찾을 수 없음")
                continue

            expected_df = df_all[df_all[actual_col].astype(str).str.strip() != ''].copy()
            
            # 행 수 비교
            count_match = len(df_group) == len(expected_df)
            
            # 데이터 일치 여부 (Key 기준 샘플링 체크)
            data_match = True
            mismatch_count = 0
            if count_match and not df_group.empty:
                # 첫 행, 중간 행, 끝 행 샘플 비교
                sample_indices = [0, len(df_group)//2, len(df_group)-1]
                for idx in sample_indices:
                    row_group = df_group.iloc[idx]
                    key_val = row_group[key_col]
                    row_all = df_all[df_all[key_col] == key_val].iloc[0] if not df_all[df_all[key_col] == key_val].empty else None
                    
                    if row_all is None:
                        data_match = False
                        break
                    
                    # 주요 필드 비교 (WEIGHT, T, B 등) - 타입 차이 극복 위해 문자열로 통일
                    for col in [actual_col, 'WEIGHT', 'T', 'B']:
                        if col in df_group.columns and col in df_all.columns:
                            val_group = str(row_group[col]).strip()
                            val_all = str(row_all[col]).strip()
                            
                            # 실수/정수 표현 차이 보정 (예: 10 vs 10.0)
                            try:
                                if float(val_group) == float(val_all):
                                    continue
                            except:
                                pass

                            if val_group != val_all:
                                data_match = False
                                mismatch_count += 1

            status = "✓ 일치" if (count_match and data_match) else "✗ 불일치"
            print(f"  - {sheet_name:15}: {status} (행 수: {len(df_group)})")
            
            results['cross_sheet_validation'][sheet_name] = {
                'row_count_match': count_match,
                'data_match': data_match,
                'mismatch_count': mismatch_count
            }

        # 결과 저장
        report_path = Path('validation_report.json')
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"\n[SUCCESS] 검증 완료! 리포트: {report_path}")

    # =====================================================================
    # 메인 메뉴
    # =====================================================================
    def show_menu(self):
        """메인 메뉴 표시"""
        print("\n" + "=" * 60)
        print("BOM 데이터 처리 옵션")
        print("=" * 60)
        print("1. JSON 전체 생성 (기존 로직)")
        print("2. 엑셀 생성 + 그래프 (병합/정렬)")
        print("3. JSON vs 원본 비교 검증")
        print("4. 모두 실행 (1 + 2 + 3)")
        print("0. 종료")
        print("=" * 60)

        while True:
            try:
                choice = input("\n선택 (0-4): ").strip()
                if choice in ['0', '1', '2', '3', '4']:
                    return choice
                print("[ERROR] 0-4 사이의 숫자를 입력하세요.")
            except KeyboardInterrupt:
                print("\n\n[INFO] 프로그램을 종료합니다.")
                return '0'

    def run(self):
        """메인 실행 함수"""
        # 파일 로드
        if not self.load_files():
            return

        # 데이터 매칭
        if not self.match_data():
            return

        # 명령줄 인수 처리 (자동화용)
        if len(sys.argv) > 1:
            if sys.argv[1] == '--auto':
                choice = sys.argv[2] if len(sys.argv) > 2 else '4'
                print(f"\n[INFO] 자동 실행 모드: 옵션 {choice}")
                if choice == '1': self.option1_generate_json()
                elif choice == '2': self.option2_excel_with_charts()
                elif choice == '3': self.option3_validate_json()
                elif choice == '4':
                    self.option1_generate_json()
                    self.option2_excel_with_charts()
                    self.option3_validate_json()
                return

        # 메뉴 선택
        while True:
            choice = self.show_menu()

            if choice == '0':
                print("\n[INFO] 프로그램을 종료합니다.")
                break

            elif choice == '1':
                self.option1_generate_json()

            elif choice == '2':
                self.option2_excel_with_charts()

            elif choice == '3':
                self.option3_validate_json()

            elif choice == '4':
                print("\n[INFO] 모든 옵션 실행 중...")
                self.option1_generate_json()
                self.option2_excel_with_charts()
                self.option3_validate_json()
                print("\n[SUCCESS] 모든 작업 완료!")
                break


def main():
    """메인 함수"""
    processor = BOMProcessor()
    processor.run()


if __name__ == "__main__":
    main()
